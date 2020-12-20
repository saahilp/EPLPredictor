import asyncio
import aiohttp
import xlrd

from understat import Understat
from openpyxl import load_workbook


# get results from Understat

async def load_results():
    async with aiohttp.ClientSession() as session:
        understat = Understat(session)
        # call to get results for 20/21
        results = await understat.get_league_results(
            "epl", 2020,
        )
        return results


# checking for duplicates in existing results file

def check_duplicate(result_to_check, sheet):
    # loop through first and second column of each row to look for duplicates
    for row in range(1, sheet.nrows):
        home_team = sheet.cell(row, 0).value
        away_team = sheet.cell(row, 1).value
        if result_to_check['h']['title'] == home_team and result_to_check['a']['title'] == away_team:
            return True
    return False


if __name__ == '__main__':

    # running async function defined above
    loop = asyncio.get_event_loop()
    league_results = loop.run_until_complete(load_results())

    # initializing library used to write to results file
    book_for_writing = load_workbook('PLResults.xlsx')
    sheet_for_writing = book_for_writing.active

    # initializing library used to read from results file
    book_for_reading = xlrd.open_workbook('PLResults.xlsx')
    sheet_for_reading = book_for_reading.sheet_by_index(0)

    # append all new results to results file
    for result in league_results:
        if not check_duplicate(result, sheet_for_reading):
            sheet_for_writing.append((result['h']['title'], result['a']['title'], result['xG']['h'], result['xG']['a']))

    book_for_writing.save('PLResults.xlsx')
